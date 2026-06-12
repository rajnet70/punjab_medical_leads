import os
import re
import time
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from bs4 import BeautifulSoup

# ── CONFIG ────────────────────────────────────────────────────────────────────
API_KEY = os.environ.get('GOOGLE_API_KEY', '')
LIMIT   = int(os.environ.get('LIMIT', '25'))
STATE   = 'Texas'

SEARCH_QUERIES = [
    'dental equipment supplier Texas',
    'dental supply company Texas',
    'dental equipment distributor Texas',
    'orthodontic supply company Texas',
    'dental software company Texas',
    'dental financing company Texas',
]

PLACES_SEARCH_URL = 'https://maps.googleapis.com/maps/api/place/textsearch/json'
PLACES_DETAIL_URL = 'https://maps.googleapis.com/maps/api/place/details/json'

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml;q=0.9,*/*;q=0.8',
    'Accept-Language': 'en-US,en;q=0.5',
}

TIMEOUT = 12

# ── STYLES ────────────────────────────────────────────────────────────────────
THIN      = Side(style='thin', color='C8D8EE')
BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FILL  = PatternFill('solid', start_color='1B3A6B')
SUB_FILL  = PatternFill('solid', start_color='2E5FA3')
ALT_FILL  = PatternFill('solid', start_color='EDF4FF')
GOLD_FILL = PatternFill('solid', start_color='FFF2CC')
ENR_FILL  = PatternFill('solid', start_color='EEE8FF')
ENR_FILL2 = PatternFill('solid', start_color='E6DEFF')

# ── SCRAPER ───────────────────────────────────────────────────────────────────
def search_places(query, page_token=None):
    params = {'query': query, 'key': API_KEY}
    if page_token:
        params['pagetoken'] = page_token
    return requests.get(PLACES_SEARCH_URL, params=params, timeout=30).json()

def get_place_details(place_id):
    params = {
        'place_id': place_id,
        'fields': 'name,formatted_phone_number,formatted_address,website,geometry',
        'key': API_KEY,
    }
    return requests.get(PLACES_DETAIL_URL, params=params, timeout=30).json().get('result', {})

def scrape_suppliers():
    seen_ids = set()
    suppliers = []

    for query in SEARCH_QUERIES:
        if len(suppliers) >= LIMIT:
            break
        print(f'Searching: {query}')
        data = search_places(query)
        for place in data.get('results', []):
            if len(suppliers) >= LIMIT:
                break
            pid = place.get('place_id')
            if not pid or pid in seen_ids:
                continue
            seen_ids.add(pid)

            print(f'  Getting details: {place.get("name")}')
            details = get_place_details(pid)
            time.sleep(0.15)

            name    = details.get('name', place.get('name', 'N/A'))
            phone   = details.get('formatted_phone_number', 'N/A') or 'N/A'
            address = details.get('formatted_address', 'N/A') or 'N/A'
            website = details.get('website', 'N/A') or 'N/A'
            geo     = details.get('geometry', {}).get('location', {})
            lat     = geo.get('lat', 'N/A')
            lng     = geo.get('lng', 'N/A')

            # Extract city from address
            parts = [p.strip() for p in address.split(',')]
            city = parts[-3] if len(parts) >= 3 else STATE

            suppliers.append({
                'name':    name,
                'city':    city,
                'phone':   phone,
                'address': address,
                'website': website,
                'lat':     lat,
                'lng':     lng,
            })

    print(f'\nScraped {len(suppliers)} suppliers')
    return suppliers[:LIMIT]

# ── ENRICHER ──────────────────────────────────────────────────────────────────
def fetch_website(url):
    if not url or url == 'N/A':
        return None
    try:
        clean = url if url.startswith('http') else f'https://{url}'
        clean = clean.split('?')[0]
        resp = requests.get(clean, headers=HEADERS, timeout=TIMEOUT, allow_redirects=True)
        if resp.status_code == 200:
            return resp.text
    except requests.exceptions.SSLError:
        try:
            resp = requests.get(url.replace('https://', 'http://'), headers=HEADERS, timeout=TIMEOUT)
            return resp.text
        except:
            return None
    except:
        return None
    return None

def extract_email(text, html):
    # From text
    emails = re.findall(r'[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}', text)
    # Filter out common false positives
    blocked = ['example', 'domain', 'email', 'user', 'test', 'placeholder', 'png', 'jpg', 'gif', 'svg']
    for e in emails:
        if not any(b in e.lower() for b in blocked):
            return e
    return 'N/A'

def extract_decision_maker(soup, text):
    name  = 'N/A'
    title = 'N/A'

    # Title patterns
    title_patterns = [
        r'(CEO|President|Owner|Founder|Director|VP|Vice President|Manager|Principal)',
    ]

    # Look for name + title combinations near contact/about sections
    about_sections = soup.find_all(['div', 'section', 'p'], class_=re.compile(r'about|team|contact|staff|leadership', re.I))
    for section in about_sections:
        section_text = section.get_text(separator=' ', strip=True)
        for pat in title_patterns:
            m = re.search(rf'([A-Z][a-z]+ [A-Z][a-z]+)[,\s]+({pat})', section_text)
            if m:
                name  = m.group(1)
                title = m.group(2)
                return name, title

    # Fallback — search full text
    for pat in title_patterns:
        m = re.search(rf'([A-Z][a-z]+ [A-Z][a-z]+)[,\s]+({pat})', text)
        if m:
            name  = m.group(1)
            title = m.group(2)
            break

    return name, title

def extract_what_they_sell(text):
    product_map = {
        'Dental Equipment':     ['dental equipment', 'dental handpiece', 'dental chair', 'dental unit', 'x-ray equipment', 'imaging equipment'],
        'Dental Supplies':      ['dental supplies', 'dental consumables', 'disposables', 'dental materials'],
        'Orthodontic Supplies': ['orthodontic', 'braces', 'aligners', 'brackets', 'orthodontic supplies'],
        'Dental Software':      ['dental software', 'practice management', 'dental imaging software', 'patient management'],
        'Dental Financing':     ['dental financing', 'patient financing', 'payment plans', 'dental loans'],
        'Sterilisation':        ['sterilization', 'sterilisation', 'autoclave', 'infection control'],
        'Dental Furniture':     ['dental furniture', 'dental cabinetry', 'operatory'],
        'Digital Dentistry':    ['digital dentistry', 'cad/cam', 'cerec', '3d printing', 'intraoral scanner'],
        'Repair & Service':     ['repair service', 'equipment repair', 'maintenance', 'servicing dental'],
    }
    found = [p for p, kws in product_map.items() if any(kw in text.lower() for kw in kws)]
    return ', '.join(found[:4]) if found else 'N/A'

def extract_territory(text):
    territory_map = {
        'Texas':         ['texas', 'tx', 'dallas', 'houston', 'austin', 'san antonio'],
        'Southwest':     ['southwest', 'new mexico', 'arizona', 'nevada'],
        'Southeast':     ['southeast', 'florida', 'georgia', 'alabama', 'louisiana'],
        'Nationwide':    ['nationwide', 'national', 'across the us', 'all 50 states', 'united states'],
        'Gulf Coast':    ['gulf coast', 'gulf region'],
        'South Central': ['south central', 'oklahoma', 'arkansas', 'mississippi'],
    }
    found = [t for t, kws in territory_map.items() if any(kw in text.lower() for kw in kws)]
    return ', '.join(found[:3]) if found else 'N/A'

def extract_brands(text):
    brand_list = [
        'Dentsply', 'Patterson', 'Henry Schein', 'Sirona', 'KaVo', 'Planmeca',
        'Carestream', 'Dexis', 'Eaglesoft', 'Dentrix', 'Ortho2', 'Dolphin',
        '3M', 'Ivoclar', 'GC America', 'Ultradent', 'Zimmer Biomet', 'Nobel Biocare',
        'Straumann', 'Ormco', '3Shape', 'Align Technology', 'Invisalign',
        'Air Techniques', 'Midmark', 'A-dec', 'Belmont', 'Pelton & Crane',
    ]
    found = [b for b in brand_list if b.lower() in text.lower()]
    return ', '.join(found[:5]) if found else 'N/A'

def extract_years(text):
    patterns = [
        r'since (\d{4})',
        r'established in (\d{4})',
        r'founded in (\d{4})',
        r'serving.*since (\d{4})',
        r'in business since (\d{4})',
        r'(\d{4}).*serving',
    ]
    for pat in patterns:
        m = re.search(pat, text.lower())
        if m:
            year = int(m.group(1))
            if 1950 <= year <= 2024:
                return f'Since {year}'
    return 'N/A'

def extract_linkedin(html):
    m = re.search(r'linkedin\.com/company/([a-zA-Z0-9\-]+)', html)
    return f'linkedin.com/company/{m.group(1)}' if m else 'N/A'

def extract_social(html):
    social = []
    if 'facebook.com' in html:   social.append('Facebook')
    if 'instagram.com' in html:  social.append('Instagram')
    if 'twitter.com' in html or 'x.com' in html: social.append('Twitter/X')
    if 'youtube.com' in html:    social.append('YouTube')
    return ', '.join(social) if social else 'N/A'

def enrich(supplier):
    print(f'  Enriching: {supplier["name"]}')
    html = fetch_website(supplier['website'])

    if not html:
        return {
            'email': 'N/A', 'dm_name': 'N/A', 'dm_title': 'N/A',
            'what_they_sell': 'N/A', 'territory': 'N/A',
            'brands': 'N/A', 'years': 'N/A',
            'linkedin': 'N/A', 'social': 'N/A',
        }

    soup = BeautifulSoup(html, 'lxml')
    text = soup.get_text(separator=' ', strip=True)
    dm_name, dm_title = extract_decision_maker(soup, text)

    return {
        'email':          extract_email(text, html),
        'dm_name':        dm_name,
        'dm_title':       dm_title,
        'what_they_sell': extract_what_they_sell(text),
        'territory':      extract_territory(text),
        'brands':         extract_brands(text),
        'years':          extract_years(text),
        'linkedin':       extract_linkedin(html),
        'social':         extract_social(html),
    }

# ── EXCEL ─────────────────────────────────────────────────────────────────────
def build_excel(data):
    os.makedirs('output', exist_ok=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    email_count = sum(1 for d in data if d['enriched']['email'] != 'N/A')
    email_rate  = f'{int(email_count/len(data)*100)}%' if data else 'N/A'

    # ── SUMMARY ──────────────────────────────────────────────────────────────
    ws = wb.create_sheet('Summary', 0)
    ws.sheet_view.showGridLines = False

    ws.merge_cells('A1:G1')
    c = ws['A1']
    c.value = 'TEXAS DENTAL SUPPLIERS — INTERNAL PROSPECTING LIST  |  LEADFLOW'
    c.font = Font(bold=True, name='Arial', size=14, color='FFFFFF')
    c.fill = HDR_FILL
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 32

    ws.merge_cells('A2:G2')
    c = ws['A2']
    c.value = '⚠  Internal Use Only — Not for Client Distribution  |  LeadFlow  |  June 2026'
    c.font = Font(bold=True, name='Arial', size=10, color='7A5C00')
    c.fill = GOLD_FILL
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 8

    hdrs   = ['State', 'Total Companies', 'Email Hit Rate', 'Category', 'Use', 'Built By', 'Date']
    widths = [14, 16, 16, 20, 24, 14, 14]
    for ci, (h, w) in enumerate(zip(hdrs, widths), 1):
        c = ws.cell(4, ci, h)
        c.font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
        c.fill = SUB_FILL
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[4].height = 20

    for ci, v in enumerate([STATE, len(data), email_rate, 'Dental Suppliers', 'Cold Email Outreach', 'LeadFlow', 'June 2026'], 1):
        c = ws.cell(5, ci, v)
        c.font = Font(name='Arial', size=10)
        c.fill = ALT_FILL
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = BORDER
    ws.row_dimensions[5].height = 20

    # ── DATA SHEET ────────────────────────────────────────────────────────────
    wd = wb.create_sheet('Texas Suppliers')
    wd.sheet_view.showGridLines = False

    headers = [
        '#', 'Company Name', 'City', 'Phone', 'Website',
        'Email *', 'Decision Maker *', 'Title *',
        'What They Sell *', 'Territory *', 'Brands Carried *',
        'Years in Business *', 'LinkedIn *', 'Social Media *'
    ]
    col_widths = [5, 36, 16, 16, 28, 28, 22, 20, 36, 24, 36, 16, 32, 24]

    # Title
    wd.merge_cells(f'A1:{get_column_letter(len(headers))}1')
    c = wd['A1']
    c.value = f'Texas Dental Suppliers — Internal Prospecting List  |  {len(data)} Companies  |  ⚠ Internal Use Only  |  LeadFlow  |  June 2026'
    c.font = Font(bold=True, name='Arial', size=10, color='FFFFFF')
    c.fill = HDR_FILL
    c.alignment = Alignment(horizontal='left', vertical='center')
    wd.row_dimensions[1].height = 22

    # Enrichment note
    wd.merge_cells(f'A2:{get_column_letter(len(headers))}2')
    c = wd['A2']
    c.value = '* Fields marked with asterisk are extracted via website analysis. N/A shown where website unavailable or data not detected.'
    c.font = Font(italic=True, name='Arial', size=9, color='5B3FA8')
    c.fill = PatternFill('solid', start_color='F3EEFF')
    c.alignment = Alignment(horizontal='left', vertical='center')
    wd.row_dimensions[2].height = 16

    # Headers
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = wd.cell(3, ci, h)
        if '*' in str(h):
            c.font = Font(bold=True, color='5B3FA8', name='Arial', size=10)
            c.fill = PatternFill('solid', start_color='DDD0FF')
        else:
            c.font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
            c.fill = SUB_FILL
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = BORDER
        wd.column_dimensions[get_column_letter(ci)].width = w
    wd.row_dimensions[3].height = 22

    # Data
    for ri, d in enumerate(data, 4):
        rfill = ALT_FILL if ri % 2 == 0 else PatternFill()
        e     = d['enriched']

        vals = [
            ri - 3,
            d['name'],
            d['city'],
            d['phone'],
            str(d['website']).split('?')[0],
            e['email'],
            e['dm_name'],
            e['dm_title'],
            e['what_they_sell'],
            e['territory'],
            e['brands'],
            e['years'],
            e['linkedin'],
            e['social'],
        ]

        for ci, v in enumerate(vals, 1):
            c = wd.cell(ri, ci, v)
            c.font = Font(name='Arial', size=10, color='000000')
            if ci >= 6:
                c.fill = ENR_FILL if ri % 2 == 0 else ENR_FILL2
            else:
                c.fill = rfill
            c.alignment = Alignment(vertical='center', wrap_text=True)
            c.border = BORDER
        wd.row_dimensions[ri].height = 28

    wd.freeze_panes = 'A4'

    path = 'output/Texas_Dental_Suppliers.xlsx'
    wb.save(path)
    print(f'\n✅ Saved: {path}')
    print(f'   Companies: {len(data)}')
    print(f'   Email hit rate: {email_rate}')

# ── MAIN ──────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    print(f'🦷 Texas Dental Supplier Scraper + Enricher')
    print(f'   Limit: {LIMIT}')
    print(f'   API Key: {"✅ Found" if API_KEY else "❌ Missing"}')

    if not API_KEY:
        print('ERROR: GOOGLE_API_KEY not set')
        exit(1)

    suppliers = scrape_suppliers()

    print(f'\n🔬 Enriching {len(suppliers)} websites...')
    enriched_data = []
    for s in suppliers:
        enriched = enrich(s)
        enriched_data.append({**s, 'enriched': enriched})
        time.sleep(0.3)

    build_excel(enriched_data)
    print('\n✅ Done!')
