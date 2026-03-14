import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import time
import os
import logging

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger(__name__)

API_KEY = os.environ.get("GOOGLE_API_KEY", "")
OUTPUT_DIR = "output"

FAISALABAD_ZONES = [
    {"name": "City Center",    "lat": 31.4180, "lng": 73.0790},
    {"name": "Clock Tower",    "lat": 31.4154, "lng": 73.0839},
    {"name": "Peoples Colony", "lat": 31.4350, "lng": 73.0780},
    {"name": "Gulberg",        "lat": 31.4450, "lng": 73.0900},
    {"name": "Madina Town",    "lat": 31.4280, "lng": 73.1050},
    {"name": "Johar Town",     "lat": 31.4050, "lng": 73.0950},
    {"name": "Jinnah Colony",  "lat": 31.4000, "lng": 73.0700},
    {"name": "Samanabad",      "lat": 31.4300, "lng": 73.0600},
    {"name": "Canal Road",     "lat": 31.4500, "lng": 73.1100},
    {"name": "Satiana Road",   "lat": 31.3900, "lng": 73.0800},
    {"name": "Sargodha Road",  "lat": 31.4600, "lng": 73.0700},
    {"name": "Jaranwala Road", "lat": 31.4100, "lng": 73.1200},
    {"name": "Millat Road",    "lat": 31.4400, "lng": 73.1300},
    {"name": "Susan Road",     "lat": 31.4650, "lng": 73.1000},
    {"name": "Dijkot Road",    "lat": 31.3800, "lng": 73.0600},
]

DOCTOR_QUERIES = ["doctor", "clinic", "hospital", "specialist", "medical center"]
STORE_QUERIES = ["pharmacy", "medical store", "dawakhana", "chemist"]
RADIUS = 1500


def nearby_search(lat, lng, keyword):
    results = []
    url = "https://maps.googleapis.com/maps/api/place/nearbysearch/json"
    params = {
        "location": f"{lat},{lng}",
        "radius": RADIUS,
        "keyword": keyword,
        "key": API_KEY
    }
    for page in range(3):
        try:
            r = requests.get(url, params=params, timeout=15)
            data = r.json()
            status = data.get("status")
            if status == "REQUEST_DENIED":
                log.error(f"API key error: {data.get('error_message')}")
                return results
            if status not in ("OK", "ZERO_RESULTS"):
                break
            results.extend(data.get("results", []))
            next_token = data.get("next_page_token")
            if not next_token:
                break
            time.sleep(2)
            params = {"pagetoken": next_token, "key": API_KEY}
        except Exception as e:
            log.warning(f"Search error: {e}")
            break
    return results


def get_phone(place_id):
    try:
        r = requests.get(
            "https://maps.googleapis.com/maps/api/place/details/json",
            params={
                "place_id": place_id,
                "fields": "formatted_phone_number,international_phone_number",
                "key": API_KEY
            },
            timeout=10
        )
        result = r.json().get("result", {})
        return result.get("formatted_phone_number") or result.get("international_phone_number") or ""
    except:
        return ""


def dedup(places):
    seen, out = set(), []
    for p in places:
        pid = p.get("place_id")
        if pid and pid not in seen:
            seen.add(pid)
            out.append(p)
    return out


def collect(zones, queries, label):
    all_places = []
    for zone in zones:
        log.info(f"  [{label}] Zone: {zone['name']}")
        for query in queries:
            places = nearby_search(zone["lat"], zone["lng"], query)
            all_places.extend(places)
            log.info(f"    {query}: {len(places)} results")
            time.sleep(0.5)
    all_places = dedup(all_places)
    log.info(f"  [{label}] {len(all_places)} unique after dedup")
    results = []
    total = len(all_places)
    for i, place in enumerate(all_places):
        name = place.get("name", "")
        address = place.get("vicinity", "") or place.get("formatted_address", "")
        pid = place.get("place_id", "")
        rating = str(place.get("rating", ""))
        phone = get_phone(pid) if pid else ""
        time.sleep(0.3)
        results.append({
            "name": name,
            "address": address,
            "phone": phone,
            "rating": rating,
        })
        if (i + 1) % 10 == 0:
            log.info(f"  [{label}] Fetched phones: {i+1}/{total}")
    return results


BORDER = Border(
    left=Side(style="thin", color="C8D8EE"),
    right=Side(style="thin", color="C8D8EE"),
    top=Side(style="thin", color="C8D8EE"),
    bottom=Side(style="thin", color="C8D8EE"),
)


def write_sheet(ws, title_text, rows, cols, widths, color, val_fn):
    c = ws.cell(row=1, column=1, value=f"{title_text}  |  Total: {len(rows)}")
    c.font = Font(bold=True, name="Calibri", size=13, color="1B3A6B")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
    ws.row_dimensions[1].height = 26
    fill = PatternFill("solid", start_color=color)
    font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    for ci, (h, w) in enumerate(zip(cols, widths), 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = font
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 22
    alt = PatternFill("solid", start_color="EDF4FF")
    for ri, row in enumerate(rows, 3):
        rfill = alt if ri % 2 == 0 else PatternFill()
        for ci, v in enumerate(val_fn(ri - 2, row), 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font = Font(name="Calibri", size=10)
            c.fill = rfill
            c.alignment = Alignment(vertical="center", wrap_text=True)
            c.border = BORDER
        ws.row_dimensions[ri].height = 18
    ws.freeze_panes = "A3"


def make_excel(city, doctors, stores):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws1 = wb.create_sheet("Doctors")
    write_sheet(
        ws1,
        f"{city} — Doctors",
        doctors,
        ["#", "Name", "Phone", "Address", "Rating"],
        [5, 35, 18, 50, 8],
        "1B3A6B",
        lambda i, d: [i, d["name"], d["phone"], d["address"], d["rating"]]
    )
    ws2 = wb.create_sheet("Medical Stores")
    write_sheet(
        ws2,
        f"{city} — Medical Stores & Pharmacies",
        stores,
        ["#", "Store Name", "Phone", "Address", "Rating"],
        [5, 35, 18, 50, 8],
        "1A5C3A",
        lambda i, s: [i, s["name"], s["phone"], s["address"], s["rating"]]
    )
    path = os.path.join(OUTPUT_DIR, f"{city}.xlsx")
    wb.save(path)
    log.info(f"Saved: {path} ({len(doctors)} doctors, {len(stores)} stores)")


def main():
    if not API_KEY:
        log.error("GOOGLE_API_KEY not set! Add it as a GitHub secret.")
        exit(1)
    city = "Faisalabad"
    log.info(f"Starting grid scrape for {city}")
    log.info(f"Zones: {len(FAISALABAD_ZONES)}")
    log.info("Collecting doctors...")
    doctors = collect(FAISALABAD_ZONES, DOCTOR_QUERIES, "Doctors")
    log.info("Collecting medical stores...")
    stores = collect(FAISALABAD_ZONES, STORE_QUERIES, "Stores")
    log.info(f"Final — Doctors: {len(doctors)} | Stores: {len(stores)}")
    make_excel(city, doctors, stores)
    log.info("All done!")


if __name__ == "__main__":
    main()
