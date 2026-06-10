import os
import time
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── CONFIG ────────────────────────────────────────────────────────────────────
API_KEY = os.environ.get('GOOGLE_API_KEY', '')
CITY    = os.environ.get('CITY', 'New York')
LIMIT   = int(os.environ.get('LIMIT', '50'))

SEARCH_QUERIES = [
    f'dental clinic {CITY}',
    f'dentist {CITY}',
    f'dental office {CITY}',
    f'orthodontist {CITY}',
    f'dental surgery {CITY}',
]

PLACES_SEARCH_URL = 'https://maps.googleapis.com/maps/api/place/textsearch/json'
PLACES_DETAIL_URL = 'https://maps.googleapis.com/maps/api/place/details/json'

# ── STYLES ────────────────────────────────────────────────────────────────────
BORDER = Border(
    left=Side(style='thin', color='C8D8EE'),
    right=Side(style='thin', color='C8D8EE'),
    top=Side(style='thin', color='C8D8EE'),
    bottom=Side(style='thin', color='C8D8EE'),
)
HEADER_FILL  = PatternFill('solid', start_color='1B3A6B')
SUBHEAD_FILL = PatternFill('solid', start_color='2E5FA3')
ALT_FILL     = PatternFill('solid', start_color='EDF4FF')
HIGH_FILL    = PatternFill('solid', start_color='C6EFCE')
MED_FILL     = PatternFill('solid', start_color='FFEB9C')
LOW_FILL     = PatternFill('solid', start_color='FFC7CE')
GOLD_FILL    = PatternFill('solid', start_color='FFF2CC')

# ── SCRAPER ───────────────────────────────────────────────────────────────────
def search_places(query, page_token=None):
    params = {
        'query': query,
        'key': API_KEY,
        'type': 'dentist',
    }
    if page_token:
        params['pagetoken'] = page_token
    resp = requests.get(PLACES_SEARCH_URL, params=params, timeout=30)
    return resp.json()

def get_place_details(place_id):
    params = {
        'place_id': place_id,
        'fields': 'name,formatted_phone_number,formatted_address,website,rating,user_ratings_total,geometry,opening_hours,types',
        'key': API_KEY,
    }
    resp = requests.get(PLACES_DETAIL_URL, params=params, timeout=30)
    return resp.json().get('result', {})

def extract_area(address):
    parts = address.split(',')
    if len(parts) >= 2:
        return parts[-3].strip() if len(parts) >= 3 else parts[-2].strip()
    return address

def score_clinic(clinic):
    score = 0
    rating = clinic.get('rating', 0) or 0
    reviews = clinic.get('reviews', 0) or 0
    website = clinic.get('website', '') or ''
    hours   = clinic.get('hours', '') or ''

    if rating >= 4.5: score += 25
    elif rating >= 4.0: score += 15
    elif rating >= 3.5: score += 8

    if reviews >= 200: score += 25
    elif reviews >= 100: score += 18
    elif reviews >= 50: score += 12
    elif reviews >= 20: score += 6

    if website and website != 'N/A': score += 20

    if hours and hours != 'N/A': score += 10

    if rating >= 4.0 and reviews >= 50: score += 10
    if rating >= 4.5 and reviews >= 100: score += 10

    return min(score, 100)

def get_tier(score):
    if score >= 65: return 'High'
    if score >= 40: return 'Medium'
    return 'Low'

def scrape_clinics():
    seen_ids = set()
    clinics  = []

    for query in SEARCH_QUERIES:
        if len(clinics) >= LIMIT:
            break

        print(f'Searching: {query}')
        page_token = None

        for _ in range(3):
            if len(clinics) >= LIMIT:
                break

            data = search_places(query, page_token)
            results = data.get('results', [])

            for place in results:
                if len(clinics) >= LIMIT:
                    break

                pid = place.get('place_id')
                if not pid or pid in seen_ids:
                    continue
                seen_ids.add(pid)

                print(f'  Getting details for: {place.get("name")}')
                details = get_place_details(pid)
                time.sleep(0.15)

                name    = details.get('name', place.get('name', 'N/A'))
                phone   = details.get('formatted_phone_number', 'N/A') or 'N/A'
                address = details.get('formatted_address', place.get('formatted_address', 'N/A'))
                website = details.get('website', 'N/A') or 'N/A'
                rating  = details.get('rating', place.get('rating', 0)) or 0
                reviews = details.get('user_ratings_total', place.get('user_ratings_total', 0)) or 0
                geo     = details.get('geometry', {}).get('location', {})
                lat     = geo.get('lat', 'N/A')
                lng     = geo.get('lng', 'N/A')

                hours_list = details.get('opening_hours', {}).get('weekday_text', [])
                hours = hours_list[0] if hours_list else 'N/A'

                area = extract_area(address)

                clinic = {
                    'name':    name,
                    'phone':   phone,
                    'address': address,
                    'area':    area,
                    'website': website,
                    'rating':  round(float(rating), 1),
                    'reviews': int(reviews),
                    'lat':     lat,
                    'lng':     lng,
                    'hours':   hours,
                }
                clinic['score'] = score_clinic(clinic)
                clinic['tier']  = get_tier(clinic['score'])
                clinics.append(clinic)

            next_token = data.get('next_page_token')
            if not next_token:
                break
            page_token = next_token
            time.sleep(2)

    print(f'\nTotal scraped: {len(clinics)}')
    return clinics[:LIMIT]

# ── EXCEL ─────────────────────────────────────────────────────────────────────
def build_excel(clinics):
    os.makedirs('output', exist_ok=True)

    high = sum(1 for c in clinics if c['tier'] == 'High')
    med  = sum(1 for c in clinics if c['tier'] == 'Medium')
    low  = sum(1 for c in clinics if c['tier'] == 'Low')

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── SUMMARY TAB ────────────────────────────────────────────────────────────
    ws_sum = wb.create_sheet('Summary', 0)
    ws_sum.sheet_view.showGridLines = False

    ws_sum.merge_cells('A1:G1')
    c = ws_sum.cell(row=1, column=1,
        value=f'{CITY.upper()} DENTAL CLINIC INTELLIGENCE — LEADFLOW')
    c.font = Font(bold=True, name='Calibri', size=16, color='FFFFFF')
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws_sum.row_dimensions[1].height = 36

    ws_sum.merge_cells('A2:G2')
    c = ws_sum.cell(row=2, column=1,
        value=f'⚠ Sample Database — {len(clinics)} Records  |  Prepared by LeadFlow  |  June 2026  |  For Presentation Purposes Only')
    c.font = Font(bold=True, name='Calibri', size=11, color='7a5c00')
    c.fill = GOLD_FILL
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws_sum.row_dimensions[2].height = 26

    ws_sum.row_dimensions[3].height = 10

    sum_headers = ['City', 'Total Records', 'High Priority', 'Medium Priority', 'Low Priority', 'Category', 'Notes']
    sum_widths   = [16, 14, 14, 16, 14, 16, 32]
    for ci, (h, w) in enumerate(zip(sum_headers, sum_widths), 1):
        c = ws_sum.cell(row=4, column=ci, value=h)
        c.font = Font(bold=True, color='FFFFFF', name='Calibri', size=11)
        c.fill = SUBHEAD_FILL
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = BORDER
        ws_sum.column_dimensions[get_column_letter(ci)].width = w
    ws_sum.row_dimensions[4].height = 22

    vals = [CITY, len(clinics), high, med, low, 'Dental Clinics', 'Sample only — full database available on request']
    for ci, v in enumerate(vals, 1):
        c = ws_sum.cell(row=5, column=ci, value=v)
        c.font = Font(name='Calibri', size=11)
        c.fill = ALT_FILL
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = BORDER
    ws_sum.row_dimensions[5].height = 22

    for ci, v in enumerate(['SAMPLE TOTAL', len(clinics), high, med, low, '', ''], 1):
        c = ws_sum.cell(row=6, column=ci, value=v)
        c.font = Font(bold=True, color='FFFFFF', name='Calibri', size=12)
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = BORDER
    ws_sum.row_dimensions[6].height = 26

    # ── CITY TAB ───────────────────────────────────────────────────────────────
    ws = wb.create_sheet(CITY)
    ws.sheet_view.showGridLines = False

    headers = ['#', 'Clinic Name', 'Phone', 'Address', 'Area',
               'Website', 'Rating', 'Reviews', 'Opening Hours',
               'GPS Latitude', 'GPS Longitude', 'Score', 'Tier']
    widths  = [5, 38, 16, 44, 18, 28, 8, 10, 32, 14, 14, 8, 10]

    ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
    c = ws.cell(row=1, column=1,
        value=f'{CITY} — Dental Clinic Intelligence  |  {len(clinics)} Records  |  ⚠ Sample Database — LeadFlow  |  June 2026')
    c.font = Font(bold=True, name='Calibri', size=11, color='FFFFFF')
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 26

    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = Font(bold=True, color='FFFFFF', name='Calibri', size=11)
        c.fill = SUBHEAD_FILL
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 22

    for ri, clinic in enumerate(clinics, 3):
        rfill    = ALT_FILL if ri % 2 == 0 else PatternFill()
        tier_val = clinic['tier']

        vals = [
            ri - 2,
            clinic['name'],
            clinic['phone'],
            clinic['address'],
            clinic['area'],
            clinic['website'],
            clinic['rating'],
            clinic['reviews'],
            clinic['hours'],
            clinic['lat'],
            clinic['lng'],
            clinic['score'],
            clinic['tier'],
        ]

        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font = Font(name='Calibri', size=10)
            c.fill = rfill
            c.alignment = Alignment(vertical='center', wrap_text=True)
            c.border = BORDER

        tier_cell = ws.cell(row=ri, column=13)
        tier_cell.fill = HIGH_FILL if tier_val == 'High' else (MED_FILL if tier_val == 'Medium' else LOW_FILL)
        tier_cell.font = Font(name='Calibri', size=10, bold=True)
        ws.row_dimensions[ri].height = 18

    ws.freeze_panes = 'A3'

    out_path = f'output/{CITY}_Dental_Clinics.xlsx'
    wb.save(out_path)
    print(f'\n✅ Excel saved: {out_path}')
    print(f'   High: {high} | Medium: {med} | Low: {low}')

# ── MAIN ──────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    print(f'🦷 Starting US Dental Clinic Scraper')
    print(f'   City: {CITY}')
    print(f'   Limit: {LIMIT}')
    print(f'   API Key: {"✅ Found" if API_KEY else "❌ Missing"}')

    if not API_KEY:
        print('ERROR: GOOGLE_API_KEY not set')
        exit(1)

    clinics = scrape_clinics()
    build_excel(clinics)
    print('\n✅ Done!')
