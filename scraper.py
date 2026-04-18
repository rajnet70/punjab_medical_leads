import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import time
import os
import logging

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger(__name__)

API_KEY    = os.environ.get("GOOGLE_API_KEY", "")
OUTPUT_DIR = "output"
CITY       = os.environ.get("CITY", "Lahore").strip()

# ── CITY ZONE MAPS ─────────────────────────────────────────────────────────────

CITY_ZONES = {

    "Lahore": [
        {"name": "Gulberg",           "lat": 31.5120, "lng": 74.3351},
        {"name": "DHA Phase 1-2",     "lat": 31.4811, "lng": 74.4013},
        {"name": "DHA Phase 4-5",     "lat": 31.4697, "lng": 74.4197},
        {"name": "DHA Phase 6",       "lat": 31.4550, "lng": 74.4350},
        {"name": "Johar Town",        "lat": 31.4697, "lng": 74.2728},
        {"name": "Model Town",        "lat": 31.4834, "lng": 74.3274},
        {"name": "Bahria Town",       "lat": 31.3656, "lng": 74.1847},
        {"name": "Wapda Town",        "lat": 31.4418, "lng": 74.2623},
        {"name": "Iqbal Town",        "lat": 31.4976, "lng": 74.2972},
        {"name": "Garden Town",       "lat": 31.5028, "lng": 74.3238},
        {"name": "Ferozepur Road",    "lat": 31.4729, "lng": 74.3054},
        {"name": "Township",          "lat": 31.4594, "lng": 74.2805},
        {"name": "Shahdara",          "lat": 31.6185, "lng": 74.3156},
        {"name": "Raiwind Road",      "lat": 31.4197, "lng": 74.3264},
        {"name": "Anarkali",          "lat": 31.5653, "lng": 74.3144},
        {"name": "Ichra",             "lat": 31.5231, "lng": 74.2967},
        {"name": "Samanabad",         "lat": 31.5389, "lng": 74.2897},
        {"name": "Allama Iqbal Town", "lat": 31.5028, "lng": 74.2728},
        {"name": "Cavalry Ground",    "lat": 31.5384, "lng": 74.3712},
        {"name": "Cantt",             "lat": 31.5497, "lng": 74.3587},
        {"name": "Mall Road",         "lat": 31.5204, "lng": 74.3587},
        {"name": "Civil Lines",       "lat": 31.5497, "lng": 74.3280},
        {"name": "Garhi Shahu",       "lat": 31.5380, "lng": 74.3280},
        {"name": "Shadman",           "lat": 31.5250, "lng": 74.3200},
        {"name": "Gulshan-e-Ravi",    "lat": 31.5500, "lng": 74.2750},
        {"name": "Nawaz Town",        "lat": 31.5150, "lng": 74.2600},
        {"name": "Sabzazar",          "lat": 31.5350, "lng": 74.2500},
        {"name": "Kot Lakhpat",       "lat": 31.5000, "lng": 74.2700},
        {"name": "Wahdat Road",       "lat": 31.4900, "lng": 74.3150},
        {"name": "Multan Road",       "lat": 31.4500, "lng": 74.3100},
        {"name": "Valencia Town",     "lat": 31.4600, "lng": 74.3500},
        {"name": "Lake City",         "lat": 31.4200, "lng": 74.4000},
        {"name": "Thokar Niaz Baig",  "lat": 31.4050, "lng": 74.2600},
        {"name": "Harbanspura",       "lat": 31.5750, "lng": 74.2950},
        {"name": "Aziz Bhatti Town",  "lat": 31.5900, "lng": 74.3400},
        {"name": "Badami Bagh",       "lat": 31.5730, "lng": 74.3250},
        {"name": "Islampura",         "lat": 31.5600, "lng": 74.3050},
        {"name": "Mustafa Town",      "lat": 31.5050, "lng": 74.3800},
        {"name": "Bedian Road",       "lat": 31.4313, "lng": 74.4561},
        {"name": "Wahdat Colony",     "lat": 31.4850, "lng": 74.3300},
        {"name": "Bhatta Chowk",      "lat": 31.4350, "lng": 74.2900},
    ],

    "Rawalpindi": [
        {"name": "Saddar",            "lat": 33.5970, "lng": 73.0433},
        {"name": "Bahria Town Rwp",   "lat": 33.5162, "lng": 73.0699},
        {"name": "Chaklala",          "lat": 33.6100, "lng": 73.0800},
        {"name": "Commercial Market", "lat": 33.5800, "lng": 73.0600},
        {"name": "Dhoke Kala Khan",   "lat": 33.5650, "lng": 73.0300},
        {"name": "Satellite Town",    "lat": 33.6200, "lng": 73.0550},
        {"name": "Westridge",         "lat": 33.6050, "lng": 73.0250},
        {"name": "Gulzar-e-Quaid",    "lat": 33.5500, "lng": 73.1000},
        {"name": "Adiala Road",       "lat": 33.5300, "lng": 73.0200},
        {"name": "Murree Road",       "lat": 33.6300, "lng": 73.0900},
        {"name": "Raja Bazaar",       "lat": 33.5900, "lng": 73.0500},
        {"name": "Dheri Hassanabad",  "lat": 33.5750, "lng": 73.0150},
        {"name": "Taxila Road",       "lat": 33.6500, "lng": 72.9800},
        {"name": "Chakri Road",       "lat": 33.5100, "lng": 73.0500},
        {"name": "Cantt Rwp",         "lat": 33.5950, "lng": 73.0700},
    ],

    "Gujranwala": [
        {"name": "City Center",       "lat": 32.1611, "lng": 74.1883},
        {"name": "Satellite Town",    "lat": 32.1800, "lng": 74.2000},
        {"name": "Model Town",        "lat": 32.1500, "lng": 74.1700},
        {"name": "Peoples Colony",    "lat": 32.1700, "lng": 74.2100},
        {"name": "GT Road",           "lat": 32.1400, "lng": 74.1900},
        {"name": "Rahwali",           "lat": 32.2000, "lng": 74.2200},
        {"name": "Gulshan Colony",    "lat": 32.1600, "lng": 74.1600},
        {"name": "Gondlanwala Road",  "lat": 32.1300, "lng": 74.1700},
        {"name": "Wazirabad Road",    "lat": 32.1900, "lng": 74.1500},
        {"name": "Sialkot Road",      "lat": 32.1700, "lng": 74.2300},
        {"name": "Khiali Road",       "lat": 32.1450, "lng": 74.2050},
        {"name": "Trust Colony",      "lat": 32.1750, "lng": 74.1950},
    ],

    "Multan": [
        {"name": "Gulgasht Colony",   "lat": 30.2100, "lng": 71.4700},
        {"name": "Cantt Multan",      "lat": 30.1950, "lng": 71.4800},
        {"name": "Qasim Bela",        "lat": 30.1800, "lng": 71.4600},
        {"name": "Shah Rukn-e-Alam",  "lat": 30.2000, "lng": 71.4550},
        {"name": "Bosan Road",        "lat": 30.2300, "lng": 71.4400},
        {"name": "Vehari Road",       "lat": 30.1600, "lng": 71.5000},
        {"name": "Khanewal Road",     "lat": 30.2200, "lng": 71.5100},
        {"name": "Nishatabad",        "lat": 30.1900, "lng": 71.4300},
        {"name": "New Multan",        "lat": 30.2400, "lng": 71.4600},
        {"name": "Wapda Town Multan", "lat": 30.2100, "lng": 71.5200},
        {"name": "Chungi No 9",       "lat": 30.1750, "lng": 71.4750},
        {"name": "Masoom Shah Road",  "lat": 30.2050, "lng": 71.4650},
    ],

    "Sargodha": [
        {"name": "City Center",       "lat": 32.0740, "lng": 72.6711},
        {"name": "Satellite Town",    "lat": 32.0900, "lng": 72.6800},
        {"name": "University Road",   "lat": 32.0600, "lng": 72.6600},
        {"name": "Cantt Sargodha",    "lat": 32.0800, "lng": 72.6900},
        {"name": "Lahore Road",       "lat": 32.0500, "lng": 72.6500},
        {"name": "Chak 52",           "lat": 32.1000, "lng": 72.6400},
        {"name": "Sillanwali Road",   "lat": 32.0650, "lng": 72.6850},
        {"name": "Faisal Town",       "lat": 32.0850, "lng": 72.6550},
    ],

    "Sialkot": [
        {"name": "City Center",       "lat": 32.4945, "lng": 74.5229},
        {"name": "Cantt Sialkot",     "lat": 32.5100, "lng": 74.5100},
        {"name": "Allama Iqbal Road", "lat": 32.4800, "lng": 74.5300},
        {"name": "Paris Road",        "lat": 32.5000, "lng": 74.5400},
        {"name": "Wazirabad Road",    "lat": 32.4700, "lng": 74.5100},
        {"name": "Sambrial Road",     "lat": 32.5200, "lng": 74.5500},
        {"name": "Daska Road",        "lat": 32.4600, "lng": 74.5200},
        {"name": "Model Town Sialkot","lat": 32.5050, "lng": 74.5150},
    ],

    "Bahawalpur": [
        {"name": "City Center",       "lat": 29.3956, "lng": 71.6836},
        {"name": "Cantt Bwp",         "lat": 29.4100, "lng": 71.7000},
        {"name": "Model Town Bwp",    "lat": 29.3800, "lng": 71.6700},
        {"name": "Satellite Town Bwp","lat": 29.4050, "lng": 71.6600},
        {"name": "Farid Gate",        "lat": 29.3900, "lng": 71.6900},
        {"name": "Airport Road",      "lat": 29.3700, "lng": 71.7100},
        {"name": "Baghdad ul Jadeed", "lat": 29.4200, "lng": 71.6800},
    ],

    "Gujrat": [
        {"name": "City Center",       "lat": 32.5736, "lng": 74.0790},
        {"name": "Cantt Gujrat",      "lat": 32.5900, "lng": 74.0900},
        {"name": "Lalamusa Road",     "lat": 32.5600, "lng": 74.0600},
        {"name": "Model Town Gujrat", "lat": 32.5800, "lng": 74.0700},
        {"name": "GT Road Gujrat",    "lat": 32.5500, "lng": 74.0800},
        {"name": "Kharian Road",      "lat": 32.5650, "lng": 74.1000},
    ],

    "Sahiwal": [
        {"name": "City Center",       "lat": 30.6682, "lng": 73.1066},
        {"name": "Model Town Sahiwal","lat": 30.6800, "lng": 73.1200},
        {"name": "Cantt Sahiwal",     "lat": 30.6550, "lng": 73.0900},
        {"name": "Farid Town",        "lat": 30.6750, "lng": 73.0800},
        {"name": "Pakpattan Road",    "lat": 30.6500, "lng": 73.1100},
        {"name": "Faisalabad Road",   "lat": 30.6900, "lng": 73.1300},
    ],

}

DOCTOR_QUERIES = ["doctor", "clinic", "hospital", "specialist", "medical center"]
RADIUS = 1500

REMOVE_KEYWORDS = [
    "hijama", "cupping", "hakeem", "hakim", "unani", "tibb",
    "homeopath", "homoeopath", "homoeo", "homeo", "eastern medicine",
    "herbal", "dawakhana", "pansar",
    "physio", "rehab", "rehabilitation",
    "vet", "veterinar", "animal hospital", "animal clinic", "pet clinic",
    "meow", "woof",
    "salon", "spa", "beauty", "aesthetic lounge", "hair club",
    "hair solution", "korean aesthetic", "medi spa",
    "pizza", "cafe", "restaurant", "food center", "bakers",
    "sweets", "sabroso", "paratha", "chicken specialist",
    "auto", "garage", "motor", "workshop", "car ac", "alignment center",
    "silencer", "tyre", "electrician", "bike",
    "iphone", "mobile mart", "seo specialist", "study abroad",
    "amazon specialist", "pool doctor",
    "school", "academy",
    "ambulance", "ventilator", "optics", "optician",
]

REMOVE_EXACT = [
    "doctor saucy", "doctor pizza", "doctor cafe", "doctors cafe",
    "doctors hostel", "national auto works", "shafiq autos",
    "yousaf autos", "asif auto", "umair auto garage",
    "am-zone", "gm mart doctor city", "the doctor school",
    "doctor pharmacy", "doctors pharmacy",
]

SPECIALTY_MAP = [
    (["gastro", "stomach", "liver", "hepat", "hepatolog",
      "gastroenterolog", "ercp", "endoscop"],             "Gastroenterologist"),
    (["cardiac", "heart", "cardio", "cardiolog"],          "Cardiologist"),
    (["neuro", "brain", "spine surgeon", "neurosurg"],     "Neurologist"),
    (["diabetes", "diabetic", "endocrin", "thyroid",
      "hormone"],                                          "Endocrinologist"),
    (["urol", "kidney", "renal", "nephrolog"],             "Urologist"),
    (["gynae", "gyneco", "obstet", "maternity",
      "zicha", "fertility"],                               "Gynecologist"),
    (["child", "pediatric", "paediat", "children",
      "neonatolog"],                                       "Pediatrician"),
    (["eye", "ophthal", "vision", "retina"],               "Eye Specialist"),
    (["skin", "dermat", "laser clinic",
      "aesthetic clinic", "aesthetica"],                   "Dermatologist"),
    (["bone", "ortho", "trauma", "joint"],                 "Orthopedic"),
    (["dental", "dentist", "teeth", "tooth",
      "orthodon", "oral"],                                 "Dentist"),
    (["psychiat", "mental health"],                        "Psychiatrist"),
    (["surgeon", "surgery", "surgical", "plastic"],        "Surgeon"),
    (["cancer", "oncolog"],                                "Oncologist"),
    (["chest", "pulmo", "respirat", "tb "],                "Chest Specialist"),
    (["ent specialist", "ear nose throat",
      "ear, nose"],                                        "ENT Specialist"),
]


def nearby_search(lat, lng, keyword):
    results = []
    url = "https://maps.googleapis.com/maps/api/place/nearbysearch/json"
    params = {
        "location": f"{lat},{lng}",
        "radius": RADIUS,
        "keyword": keyword,
        "key": API_KEY
    }
    for _ in range(3):
        try:
            r = requests.get(url, params=params, timeout=15)
            data = r.json()
            status = data.get("status")
            if status == "REQUEST_DENIED":
                log.error(f"API key error: {data.get('error_message')}")
                return results
            if status not in ("OK", "ZERO_RESULTS"):
                break
            results.extend(data.get("results", []))
            next_token = data.get("next_page_token")
            if not next_token:
                break
            time.sleep(2)
            params = {"pagetoken": next_token, "key": API_KEY}
        except Exception as e:
            log.warning(f"Search error: {e}")
            break
    return results


def get_details(place_id):
    try:
        r = requests.get(
            "https://maps.googleapis.com/maps/api/place/details/json",
            params={
                "place_id": place_id,
                "fields": "formatted_phone_number,international_phone_number,website,opening_hours,user_ratings_total,geometry",
                "key": API_KEY
            },
            timeout=10
        )
        result = r.json().get("result", {})
        phone      = result.get("formatted_phone_number") or result.get("international_phone_number") or ""
        website    = result.get("website", "")
        hours_data = result.get("opening_hours", {})
        hours = ""
        if hours_data:
            weekday_text = hours_data.get("weekday_text", [])
            if weekday_text:
                hours = " | ".join(weekday_text[:3])
        pop_score = result.get("user_ratings_total", 0)
        location  = result.get("geometry", {}).get("location", {})
        lat = location.get("lat", "")
        lng = location.get("lng", "")
        gps = f"{lat}, {lng}" if lat and lng else ""
        return phone, website, hours, pop_score, gps
    except:
        return "", "", "", 0, ""


def dedup(places):
    seen, out = set(), []
    for p in places:
        pid = p.get("place_id")
        if pid and pid not in seen:
            seen.add(pid)
            out.append(p)
    return out


def should_remove(name, addr):
    text = (str(name or "") + " " + str(addr or "")).lower()
    name_lower = str(name or "").lower().strip()
    for exact in REMOVE_EXACT:
        if name_lower == exact or name_lower.startswith(exact):
            return True
    return any(k in text for k in REMOVE_KEYWORDS)


def get_specialty(name, addr):
    text = (str(name or "") + " " + str(addr or "")).lower()
    for keywords, spec in SPECIALTY_MAP:
        for kw in keywords:
            if kw in text:
                return spec
    return "GP"


def get_tier(rating, pop_score):
    if rating >= 4.5 and pop_score >= 100:
        return "High"
    elif rating >= 4.0 and pop_score >= 30:
        return "Medium"
    else:
        return "Low"


def get_area(address):
    parts = str(address or "").split(",")
    if len(parts) >= 3:
        return parts[-3].strip()
    elif len(parts) >= 2:
        return parts[-2].strip()
    return ""


def is_mobile(phone):
    p = str(phone or "").strip().replace(" ", "").replace("-", "")
    return p.startswith("03") or p.startswith("+923")


def safe_float(val):
    try:
        return float(val)
    except:
        return 0.0


def collect_doctors(zones):
    all_places = []
    for zone in zones:
        log.info(f"Zone: {zone['name']}")
        for query in DOCTOR_QUERIES:
            places = nearby_search(zone["lat"], zone["lng"], query)
            all_places.extend(places)
            log.info(f"  {query}: {len(places)} results")
            time.sleep(0.5)

    all_places = dedup(all_places)
    log.info(f"Unique places before filter: {len(all_places)}")

    results = []
    removed = 0
    total   = len(all_places)

    for i, place in enumerate(all_places):
        name    = place.get("name", "")
        address = place.get("vicinity", "") or place.get("formatted_address", "")

        if should_remove(name, address):
            removed += 1
            continue

        pid       = place.get("place_id", "")
        rating    = safe_float(place.get("rating", 0))
        specialty = get_specialty(name, address)
        area      = get_area(address)

        phone, website, hours, pop_score, gps = get_details(pid) if pid else ("", "", "", 0, "")
        time.sleep(0.3)

        tier   = get_tier(rating, pop_score)
        mobile = is_mobile(phone)

        results.append({
            "name":      name,
            "specialty": specialty,
            "phone":     phone,
            "mobile":    mobile,
            "address":   address,
            "area":      area,
            "rating":    rating,
            "pop_score": pop_score,
            "tier":      tier,
            "hours":     hours,
            "website":   website,
            "gps":       gps,
        })

        if (i + 1) % 20 == 0:
            log.info(f"Processed {i+1}/{total} | Removed: {removed}")

    results.sort(key=lambda x: (0 if x["mobile"] else 1, -x["rating"], -x["pop_score"]))
    log.info(f"Final: {len(results)} doctors | Removed: {removed}")
    return results


BORDER = Border(
    left=Side(style="thin",   color="C8D8EE"),
    right=Side(style="thin",  color="C8D8EE"),
    top=Side(style="thin",    color="C8D8EE"),
    bottom=Side(style="thin", color="C8D8EE"),
)


def make_cover(ws, city, total):
    for row in range(1, 14):
        for col in range(1, 9):
            ws.cell(row=row, column=col).fill = PatternFill("solid", start_color="1B3A6B")
    ws.column_dimensions["A"].width = 70
    ws.row_dimensions[2].height = 55
    ws.row_dimensions[3].height = 30

    ws.merge_cells("A2:H2")
    c = ws.cell(row=2, column=1, value=f"{city} Doctor Leads")
    c.font = Font(bold=True, name="Calibri", size=30, color="FFFFFF")
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A3:H3")
    c = ws.cell(row=3, column=1, value=f"Verified Medical Leads — {city}, Punjab, Pakistan")
    c.font = Font(name="Calibri", size=14, color="8aadd4")
    c.alignment = Alignment(horizontal="center", vertical="center")

    stats = [
        f"📍  City: {city}, Punjab, Pakistan",
        f"🩺  Full Database: {total} Verified Doctors (Filtered & Cleaned)",
        "📋  Sample Sheet: Top 50 Leads (Mobile Numbers Only, Highest Rated)",
        "⭐  Includes: Rating, Popularity Score, Tier, Specialty, Opening Hours, GPS",
        "🗂️  Specialties: GP, ENT, Dentist, Dermatologist, Pediatrician, Cardiologist & more",
        "🔄  Data refreshed monthly — LeadFlow Pakistan",
    ]
    for i, stat in enumerate(stats, 5):
        ws.merge_cells(f"A{i}:H{i}")
        c = ws.cell(row=i, column=1, value=stat)
        c.font = Font(name="Calibri", size=12, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[i].height = 24

    ws.merge_cells("A12:H12")
    c = ws.cell(row=12, column=1,
                value="LeadFlow Pakistan — AI-Powered B2B Lead Generation")
    c.font = Font(name="Calibri", size=11, color="FFD700", italic=True, bold=True)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[12].height = 30


def write_data_sheet(ws, rows, sheet_title):
    headers = ["#", "Doctor / Clinic", "Specialty", "Phone", "Area", "Address",
               "Rating", "Popularity Score", "Tier", "Opening Hours", "Website", "GPS"]
    widths  = [5, 32, 18, 16, 18, 42, 8, 16, 10, 40, 25, 22]

    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    c = ws.cell(row=1, column=1, value=sheet_title)
    c.font = Font(bold=True, name="Calibri", size=13, color="1B3A6B")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26

    fill = PatternFill("solid", start_color="1B3A6B")
    font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = font; c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 22

    alt = PatternFill("solid", start_color="EDF4FF")
    tier_colors = {"High": "C6EFCE", "Medium": "FFEB9C", "Low": "FFC7CE"}

    for ri, doc in enumerate(rows, 3):
        rfill = alt if ri % 2 == 0 else PatternFill()
        vals = [
            ri - 2, doc["name"], doc["specialty"], doc["phone"],
            doc["area"], doc["address"], doc["rating"] or "",
            doc["pop_score"] or "", doc["tier"],
            doc["hours"], doc["website"], doc["gps"]
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font = Font(name="Calibri", size=10)
            c.fill = rfill
            c.alignment = Alignment(vertical="center", wrap_text=True)
            c.border = BORDER
        tier_val = doc["tier"]
        if tier_val in tier_colors:
            ws.cell(row=ri, column=9).fill = PatternFill("solid", start_color=tier_colors[tier_val])
            ws.cell(row=ri, column=9).font = Font(name="Calibri", size=10, bold=True)
        ws.row_dimensions[ri].height = 18

    ws.freeze_panes = "A3"


def make_excel(city, doctors):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    top50 = [d for d in doctors if d["mobile"]][:50]

    ws_cover = wb.create_sheet("About This Data")
    make_cover(ws_cover, city, len(doctors))

    ws_sample = wb.create_sheet("Top 50 Sample")
    write_data_sheet(ws_sample, top50,
        f"{city} Doctors — Top 50 Sample  |  Mobile Numbers  |  Sorted by Rating")

    ws_full = wb.create_sheet("Full Doctor List")
    write_data_sheet(ws_full, doctors,
        f"{city} Doctors — Full List  |  Total: {len(doctors)}")

    path = os.path.join(OUTPUT_DIR, f"{city}_Doctors.xlsx")
    wb.save(path)
    log.info(f"Saved: {path}")


def main():
    if not API_KEY:
        log.error("GOOGLE_API_KEY not set!")
        exit(1)

    if CITY not in CITY_ZONES:
        log.error(f"City '{CITY}' not found! Available: {list(CITY_ZONES.keys())}")
        exit(1)

    zones = CITY_ZONES[CITY]
    log.info(f"Starting doctor scrape for {CITY}")
    log.info(f"Zones: {len(zones)} | Queries: {len(DOCTOR_QUERIES)}")

    doctors = collect_doctors(zones)
    make_excel(CITY, doctors)
    log.info(f"All done! {len(doctors)} doctors saved for {CITY}")


if __name__ == "__main__":
    main()
