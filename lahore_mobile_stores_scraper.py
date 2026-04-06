import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import time
import os
import logging

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger(__name__)

API_KEY = os.environ.get("GOOGLE_API_KEY", "")
OUTPUT_DIR = "output"
CITY = "Lahore"

LAHORE_ZONES = [
    {"name": "Mall Road",           "lat": 31.5204, "lng": 74.3587},
    {"name": "Gulberg",             "lat": 31.5120, "lng": 74.3351},
    {"name": "DHA Phase 1-2",       "lat": 31.4811, "lng": 74.4013},
    {"name": "DHA Phase 4-5",       "lat": 31.4697, "lng": 74.4197},
    {"name": "Johar Town",          "lat": 31.4697, "lng": 74.2728},
    {"name": "Model Town",          "lat": 31.4834, "lng": 74.3274},
    {"name": "Bahria Town",         "lat": 31.3656, "lng": 74.1847},
    {"name": "Wapda Town",          "lat": 31.4418, "lng": 74.2623},
    {"name": "Iqbal Town",          "lat": 31.4976, "lng": 74.2972},
    {"name": "Garden Town",         "lat": 31.5028, "lng": 74.3238},
    {"name": "Ferozepur Road",      "lat": 31.4729, "lng": 74.3054},
    {"name": "Township",            "lat": 31.4594, "lng": 74.2805},
    {"name": "Shahdara",            "lat": 31.6185, "lng": 74.3156},
    {"name": "Raiwind Road",        "lat": 31.4197, "lng": 74.3264},
    {"name": "Bedian Road",         "lat": 31.4313, "lng": 74.4561},
    {"name": "Anarkali",            "lat": 31.5653, "lng": 74.3144},
    {"name": "Ichra",               "lat": 31.5231, "lng": 74.2967},
    {"name": "Samanabad",           "lat": 31.5389, "lng": 74.2897},
    {"name": "Allama Iqbal Town",   "lat": 31.5028, "lng": 74.2728},
    {"name": "Cavalry Ground",      "lat": 31.5384, "lng": 74.3712},
]

MOBILE_QUERIES = [
    "mobile phone shop",
    "mobile store",
    "smartphone shop",
    "mobile accessories",
    "phone repair shop",
]

RADIUS = 1500

REMOVE_KEYWORDS = [
    "restaurant", "cafe", "food", "bakers", "sweets",
    "auto", "garage", "motor", "workshop",
    "salon", "spa", "beauty",
    "school", "academy", "university",
    "hospital", "clinic", "pharmacy",
    "property", "real estate",
]


def nearby_search(lat, lng, keyword):
    results = []
    url = "https://maps.googleapis.com/maps/api/place/nearbysearch/json"
    params = {
        "location": f"{lat},{lng}",
        "radius": RADIUS,
        "keyword": keyword,
        "key": API_KEY
    }
    for _ in range(3):
        try:
            r = requests.get(url, params=params, timeout=15)
            data = r.json()
            status = data.get("status")
            if status == "REQUEST_DENIED":
                log.error(f"API key error: {data.get('error_message')}")
                return results
            if status not in ("OK", "ZERO_RESULTS"):
                break
            results.extend(data.get("results", []))
            next_token = data.get("next_page_token")
            if not next_token:
                break
            time.sleep(2)
            params = {"pagetoken": next_token, "key": API_KEY}
        except Exception as e:
            log.warning(f"Search error: {e}")
            break
    return results


def get_details(place_id):
    try:
        r = requests.get(
            "https://maps.googleapis.com/maps/api/place/details/json",
            params={
                "place_id": place_id,
                "fields": "formatted_phone_number,international_phone_number,opening_hours,user_ratings_total,geometry",
                "key": API_KEY
            },
            timeout=10
        )
        result = r.json().get("result", {})
        phone     = result.get("formatted_phone_number") or result.get("international_phone_number") or ""
        hours_data = result.get("opening_hours", {})
        hours = ""
        if hours_data:
            weekday_text = hours_data.get("weekday_text", [])
            if weekday_text:
                hours = " | ".join(weekday_text[:3])
        pop_score = result.get("user_ratings_total", 0)
        location  = result.get("geometry", {}).get("location", {})
        lat = location.get("lat", "")
        lng = location.get("lng", "")
        return phone, hours, pop_score
    except:
        return "", "", 0


def dedup(places):
    seen, out = set(), []
    for p in places:
        pid = p.get("place_id")
        if pid and pid not in seen:
            seen.add(pid)
            out.append(p)
    return out


def should_remove(name, addr):
    text = (str(name or "") + " " + str(addr or "")).lower()
    return any(k in text for k in REMOVE_KEYWORDS)


def get_area(address):
    parts = str(address or "").split(",")
    if len(parts) >= 3:
        return parts[-3].strip()
    elif len(parts) >= 2:
        return parts[-2].strip()
    return ""


def is_mobile_number(phone):
    p = str(phone or "").strip().replace(" ", "").replace("-", "")
    return p.startswith("03") or p.startswith("+923")


def safe_float(val):
    try:
        return float(val)
    except:
        return 0.0


def collect_stores():
    all_places = []
    for zone in LAHORE_ZONES:
        log.info(f"Zone: {zone['name']}")
        for query in MOBILE_QUERIES:
            places = nearby_search(zone["lat"], zone["lng"], query)
            all_places.extend(places)
            log.info(f"  {query}: {len(places)} results")
            time.sleep(0.5)

    all_places = dedup(all_places)
    log.info(f"Unique places before filter: {len(all_places)}")

    results = []
    removed = 0
    total   = len(all_places)

    for i, place in enumerate(all_places):
        name    = place.get("name", "")
        address = place.get("vicinity", "") or place.get("formatted_address", "")

        if should_remove(name, address):
            removed += 1
            continue

        pid      = place.get("place_id", "")
        rating   = safe_float(place.get("rating", 0))
        area     = get_area(address)

        phone, hours, pop_score = get_details(pid) if pid else ("", "", 0)
        time.sleep(0.3)

        results.append({
            "name":      name,
            "phone":     phone,
            "address":   address,
            "area":      area,
            "rating":    rating,
            "pop_score": pop_score,
            "hours":     hours,
            "mobile_no": is_mobile_number(phone),
        })

        if (i + 1) % 20 == 0:
            log.info(f"Processed {i+1}/{total} | Removed: {removed}")

    # Sort: mobile numbers first, then by rating
    results.sort(key=lambda x: (0 if x["mobile_no"] else 1, -x["rating"], -x["pop_score"]))
    log.info(f"Final: {len(results)} stores | Removed: {removed}")
    return results


BORDER = Border(
    left=Side(style="thin",   color="C8D8EE"),
    right=Side(style="thin",  color="C8D8EE"),
    top=Side(style="thin",    color="C8D8EE"),
    bottom=Side(style="thin", color="C8D8EE"),
)


def make_cover(ws, total):
    for row in range(1, 14):
        for col in range(1, 9):
            ws.cell(row=row, column=col).fill = PatternFill("solid", start_color="1A3A5C")
    ws.column_dimensions["A"].width = 70
    ws.row_dimensions[2].height = 55
    ws.row_dimensions[3].height = 30

    ws.merge_cells("A2:H2")
    c = ws.cell(row=2, column=1, value="Lahore Mobile Store Leads")
    c.font = Font(bold=True, name="Calibri", size=30, color="FFFFFF")
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A3:H3")
    c = ws.cell(row=3, column=1, value="Verified Mobile Store Leads — Lahore, Punjab, Pakistan")
    c.font = Font(name="Calibri", size=14, color="8aadd4")
    c.alignment = Alignment(horizontal="center", vertical="center")

    stats = [
        "📍  City: Lahore, Punjab, Pakistan",
        f"📱  Full Database: {total} Verified Mobile Stores",
        "📋  Sample Sheet: Top 50 Leads (Mobile Numbers, Highest Rated)",
        "⭐  Includes: Rating, Area, Address, Phone & Opening Hours",
        "🗺️  Covers: DHA, Gulberg, Johar Town, Bahria, Model Town & more",
        "🔄  Data refreshed monthly",
    ]
    for i, stat in enumerate(stats, 5):
        ws.merge_cells(f"A{i}:H{i}")
        c = ws.cell(row=i, column=1, value=stat)
        c.font = Font(name="Calibri", size=12, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[i].height = 24

    ws.merge_cells("A12:H12")
    c = ws.cell(row=12, column=1,
                value="⚠️  Sample Only — Full database & all Punjab cities available on subscription")
    c.font = Font(name="Calibri", size=11, color="FFD700", italic=True, bold=True)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[12].height = 30


def write_sample_sheet(ws, rows):
    headers = ["#", "Store Name", "Phone", "Area", "Address", "Rating", "Opening Hours"]
    widths  = [5, 35, 18, 20, 45, 8, 45]

    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    c = ws.cell(row=1, column=1,
                value="Lahore Mobile Stores — Top 50 Sample  |  Mobile Numbers  |  Sorted by Rating")
    c.font = Font(bold=True, name="Calibri", size=13, color="1A3A5C")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26

    fill = PatternFill("solid", start_color="1A3A5C")
    font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = font; c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 22

    alt = PatternFill("solid", start_color="EDF4FF")
    for ri, store in enumerate(rows, 3):
        rfill = alt if ri % 2 == 0 else PatternFill()
        vals = [
            ri - 2,
            store["name"],
            store["phone"],
            store["area"],
            store["address"],
            store["rating"] or "",
            store["hours"],
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font = Font(name="Calibri", size=10)
            c.fill = rfill
            c.alignment = Alignment(vertical="center", wrap_text=True)
            c.border = BORDER
        ws.row_dimensions[ri].height = 18
    ws.freeze_panes = "A3"


def make_excel(stores):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    top50 = [s for s in stores if s["mobile_no"]][:50]

    ws_cover = wb.create_sheet("About This Data")
    make_cover(ws_cover, len(stores))

    ws_sample = wb.create_sheet("Top 50 Sample")
    write_sample_sheet(ws_sample, top50)

    path = os.path.join(OUTPUT_DIR, f"{CITY}_Mobile_Stores.xlsx")
    wb.save(path)
    log.info(f"Saved: {path} ({len(stores)} total | {len(top50)} in sample)")


def main():
    if not API_KEY:
        log.error("GOOGLE_API_KEY not set!")
        exit(1)
    log.info(f"Starting mobile store scrape for {CITY}")
    stores = collect_stores()
    make_excel(stores)
    log.info("All done!")


if __name__ == "__main__":
    main()
