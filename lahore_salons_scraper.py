import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import time
import os
import logging

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger(__name__)

API_KEY = os.environ.get("GOOGLE_API_KEY", "")
OUTPUT_DIR = "output"
CITY = "Lahore"

# 5 zones covering affluent areas — best salon density
LAHORE_ZONES = [
    {"name": "Gulberg",        "lat": 31.5120, "lng": 74.3351},
    {"name": "DHA Phase 4-5",  "lat": 31.4697, "lng": 74.4197},
    {"name": "Johar Town",     "lat": 31.4697, "lng": 74.2728},
    {"name": "Model Town",     "lat": 31.4834, "lng": 74.3274},
    {"name": "Garden Town",    "lat": 31.5028, "lng": 74.3238},
]

SALON_QUERIES = [
    "beauty salon",
    "ladies salon",
    "hair salon",
    "women salon",
    "beauty parlour",
]

RADIUS = 1500
MAX_REVIEWS_PER_SALON = 20  # reviewers per salon


def nearby_search(lat, lng, keyword):
    results = []
    url = "https://maps.googleapis.com/maps/api/place/nearbysearch/json"
    params = {
        "location": f"{lat},{lng}",
        "radius": RADIUS,
        "keyword": keyword,
        "key": API_KEY
    }
    for _ in range(3):
        try:
            r = requests.get(url, params=params, timeout=15)
            data = r.json()
            status = data.get("status")
            if status == "REQUEST_DENIED":
                log.error(f"API key error: {data.get('error_message')}")
                return results
            if status not in ("OK", "ZERO_RESULTS"):
                break
            results.extend(data.get("results", []))
            next_token = data.get("next_page_token")
            if not next_token:
                break
            time.sleep(2)
            params = {"pagetoken": next_token, "key": API_KEY}
        except Exception as e:
            log.warning(f"Search error: {e}")
            break
    return results


def get_salon_details(place_id):
    try:
        r = requests.get(
            "https://maps.googleapis.com/maps/api/place/details/json",
            params={
                "place_id": place_id,
                "fields": "name,formatted_phone_number,international_phone_number,website,opening_hours,rating,user_ratings_total,formatted_address,reviews",
                "key": API_KEY
            },
            timeout=15
        )
        result = r.json().get("result", {})
        phone     = result.get("formatted_phone_number") or result.get("international_phone_number") or ""
        website   = result.get("website", "")
        rating    = result.get("rating", "")
        address   = result.get("formatted_address", "")
        pop_score = result.get("user_ratings_total", 0)

        hours_data = result.get("opening_hours", {})
        hours = ""
        if hours_data:
            weekday_text = hours_data.get("weekday_text", [])
            if weekday_text:
                hours = " | ".join(weekday_text[:2])

        reviews = result.get("reviews", [])
        return phone, website, rating, address, pop_score, hours, reviews
    except Exception as e:
        log.warning(f"Details error: {e}")
        return "", "", "", "", 0, "", []


def dedup_by_place_id(places):
    seen, out = set(), []
    for p in places:
        pid = p.get("place_id")
        if pid and pid not in seen:
            seen.add(pid)
            out.append(p)
    return out


def is_mobile(phone):
    p = str(phone or "").strip().replace(" ", "").replace("-", "")
    return p.startswith("03") or p.startswith("+923")


def get_area(address):
    parts = str(address or "").split(",")
    if len(parts) >= 3:
        return parts[-3].strip()
    elif len(parts) >= 2:
        return parts[-2].strip()
    return ""


BORDER = Border(
    left=Side(style="thin",   color="C8D8EE"),
    right=Side(style="thin",  color="C8D8EE"),
    top=Side(style="thin",    color="C8D8EE"),
    bottom=Side(style="thin", color="C8D8EE"),
)


def write_sheet(ws, title_text, headers, widths, color, rows_data):
    # Title
    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    c = ws.cell(row=1, column=1, value=f"{title_text}  |  Total: {len(rows_data)}")
    c.font = Font(bold=True, name="Calibri", size=13, color="1B3A6B")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26

    # Headers
    fill = PatternFill("solid", start_color=color)
    font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = font; c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 22

    # Data
    alt = PatternFill("solid", start_color="EDF4FF")
    for ri, row in enumerate(rows_data, 3):
        rfill = alt if ri % 2 == 0 else PatternFill()
        for ci, v in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font = Font(name="Calibri", size=10)
            c.fill = rfill
            c.alignment = Alignment(vertical="center", wrap_text=True)
            c.border = BORDER
        ws.row_dimensions[ri].height = 18

    ws.freeze_panes = "A3"


def make_cover(ws, salon_count, reviewer_count):
    for row in range(1, 14):
        for col in range(1, 9):
            ws.cell(row=row, column=col).fill = PatternFill("solid", start_color="1B3A6B")

    ws.column_dimensions["A"].width = 70
    ws.row_dimensions[2].height = 55
    ws.row_dimensions[3].height = 30

    ws.merge_cells("A2:H2")
    c = ws.cell(row=2, column=1, value="Lahore Salon Leads — Owners & Customers")
    c.font = Font(bold=True, name="Calibri", size=28, color="FFFFFF")
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A3:H3")
    c = ws.cell(row=3, column=1, value="Beauty Salon Data — Lahore, Punjab, Pakistan")
    c.font = Font(name="Calibri", size=14, color="8aadd4")
    c.alignment = Alignment(horizontal="center", vertical="center")

    stats = [
        "📍  City: Lahore, Punjab, Pakistan",
        f"💅  Salon Owners: {salon_count} verified salons with contact details",
        f"👩  Female Reviewers: {reviewer_count} verified customers",
        "📱  Includes: Phone, WhatsApp indicator, Address, Rating, Opening Hours",
        "🎯  Use for: Facebook Custom Audiences, WhatsApp outreach, Direct sales",
        "🔄  Data refreshed monthly",
    ]
    for i, stat in enumerate(stats, 5):
        ws.merge_cells(f"A{i}:H{i}")
        c = ws.cell(row=i, column=1, value=stat)
        c.font = Font(name="Calibri", size=12, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[i].height = 24

    ws.merge_cells("A12:H12")
    c = ws.cell(row=12, column=1,
                value="⚠️  Sample Data — Full Lahore dataset available on subscription")
    c.font = Font(name="Calibri", size=11, color="FFD700", italic=True, bold=True)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[12].height = 30


def make_excel(salons, reviewers):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Cover
    ws_cover = wb.create_sheet("About This Data")
    make_cover(ws_cover, len(salons), len(reviewers))

    # Salon Owners Sheet
    ws_salons = wb.create_sheet("Salon Owners")
    salon_headers = ["#", "Salon Name", "Phone", "Mobile?", "Area", "Address",
                     "Rating", "Total Reviews", "Opening Hours", "Website"]
    salon_widths  = [5, 32, 16, 10, 18, 42, 8, 12, 40, 28]
    salon_rows = []
    for i, s in enumerate(salons, 1):
        salon_rows.append([
            i, s["name"], s["phone"],
            "Yes" if s["is_mobile"] else "No",
            s["area"], s["address"],
            s["rating"], s["pop_score"],
            s["hours"], s["website"]
        ])
    write_sheet(ws_salons,
                "Lahore Salons — Owner Data",
                salon_headers, salon_widths, "8B2252", salon_rows)

    # Reviewers Sheet
    ws_reviewers = wb.create_sheet("Female Reviewers")
    rev_headers = ["#", "Reviewer Name", "Salon Reviewed", "Area", "Rating Given",
                   "Review Text", "Total Reviews Written"]
    rev_widths  = [5, 25, 30, 18, 12, 50, 18]
    rev_rows = []
    for i, r in enumerate(reviewers, 1):
        rev_rows.append([
            i, r["reviewer_name"], r["salon_name"],
            r["salon_area"], r["rating_given"],
            r["review_text"][:100] + "..." if len(r["review_text"]) > 100 else r["review_text"],
            r["total_reviews"]
        ])
    write_sheet(ws_reviewers,
                "Lahore Salon Customers — Female Reviewers",
                rev_headers, rev_widths, "1B3A6B", rev_rows)

    path = os.path.join(OUTPUT_DIR, f"{CITY}_Salons.xlsx")
    wb.save(path)
    log.info(f"Saved: {path} ({len(salons)} salons | {len(reviewers)} reviewers)")


def main():
    if not API_KEY:
        log.error("GOOGLE_API_KEY not set!")
        exit(1)

    log.info(f"Starting Lahore Salon scrape — 5 zones")

    # Step 1 — Collect all salons
    all_places = []
    for zone in LAHORE_ZONES:
        log.info(f"Zone: {zone['name']}")
        for query in SALON_QUERIES:
            places = nearby_search(zone["lat"], zone["lng"], query)
            all_places.extend(places)
            log.info(f"  {query}: {len(places)} results")
            time.sleep(0.5)

    all_places = dedup_by_place_id(all_places)
    log.info(f"Unique salons found: {len(all_places)}")

    # Step 2 — Get details + reviews for each salon
    salons    = []
    reviewers = []
    seen_reviewers = set()
    total = len(all_places)

    for i, place in enumerate(all_places):
        pid  = place.get("place_id", "")
        name = place.get("name", "")
        log.info(f"  [{i+1}/{total}] {name}")

        phone, website, rating, address, pop_score, hours, reviews = get_salon_details(pid)
        area     = get_area(address) or place.get("vicinity", "")
        mobile   = is_mobile(phone)

        salons.append({
            "name":      name,
            "phone":     phone,
            "is_mobile": mobile,
            "area":      area,
            "address":   address,
            "rating":    rating,
            "pop_score": pop_score,
            "hours":     hours,
            "website":   website,
        })

        # Step 3 — Extract reviewers
        for review in reviews[:MAX_REVIEWS_PER_SALON]:
            reviewer_name  = review.get("author_name", "")
            rating_given   = review.get("rating", "")
            review_text    = review.get("text", "")
            total_rev      = review.get("total_ratings", 0)

            # Skip obvious male names — basic filter
            male_indicators = ["mr.", "mr ", "muhammad ", "mohammad ", "ahmed ", "ali ", "khan ",
                               "usman", "hassan", "bilal", "omar", "umar", "asad", "fahad",
                               "imran", "tariq", "sajid", "hamid", "nawaz", "zubair"]
            name_lower = reviewer_name.lower()
            is_likely_male = any(ind in name_lower for ind in male_indicators)

            if reviewer_name and not is_likely_male:
                key = reviewer_name.lower().strip()
                if key not in seen_reviewers:
                    seen_reviewers.add(key)
                    reviewers.append({
                        "reviewer_name": reviewer_name,
                        "salon_name":    name,
                        "salon_area":    area,
                        "rating_given":  rating_given,
                        "review_text":   review_text,
                        "total_reviews": total_rev,
                    })

        time.sleep(0.4)

    log.info(f"\nFinal — Salons: {len(salons)} | Female Reviewers: {len(reviewers)}")
    make_excel(salons, reviewers)
    log.info("All done!")


if __name__ == "__main__":
    main()
