import os
import re
import time
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from bs4 import BeautifulSoup

# ── CONFIG ────────────────────────────────────────────────────────────────────
CITY   = os.environ.get('CITY', 'New York')
INPUT  = f'output/{CITY}_Dental_Clinics.xlsx'
OUTPUT = f'output/{CITY}_Dental_Clinics_Enriched.xlsx'

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml;q=0.9,*/*;q=0.8',
    'Accept-Language': 'en-US,en;q=0.5',
}

TIMEOUT = 10

# ── STYLES ────────────────────────────────────────────────────────────────────
BORDER       = Border(left=Side(style='thin',color='C8D8EE'),right=Side(style='thin',color='C8D8EE'),top=Side(style='thin',color='C8D8EE'),bottom=Side(style='thin',color='C8D8EE'))
HEADER_FILL  = PatternFill('solid', start_color='1B3A6B')
SUBHEAD_FILL = PatternFill('solid', start_color='2E5FA3')
ALT_FILL     = PatternFill('solid', start_color='EDF4FF')
HIGH_FILL    = PatternFill('solid', start_color='C6EFCE')
MED_FILL     = PatternFill('solid', start_color='FFEB9C')
LOW_FILL     = PatternFill('solid', start_color='FFC7CE')
GOLD_FILL    = PatternFill('solid', start_color='FFF2CC')
ENR_FILL     = PatternFill('solid', start_color='EEE8FF')
ENR_FILL2    = PatternFill('solid', start_color='E6DEFF')

# ── FETCHERS ──────────────────────────────────────────────────────────────────
def fetch_page(url):
    try:
        resp = requests.get(url, headers=HEADERS, timeout=TIMEOUT, allow_redirects=True)
        if resp.status_code == 200:
            return resp.text
    except:
        pass
    return None

def fetch_website(url):
    if not url or url == 'N/A':
        return None
    try:
        clean = url if url.startswith('http') else f'https://{url}'
        clean = clean.split('?')[0].split('#')[0]
        resp = requests.get(clean, headers=HEADERS, timeout=TIMEOUT, allow_redirects=True)
        if resp.status_code == 200:
            return resp.text
    except requests.exceptions.SSLError:
        try:
            resp = requests.get(url.replace('https://','http://'), headers=HEADERS, timeout=TIMEOUT)
            return resp.text
        except:
            return None
    except:
        return None
    return None

def fetch_all_pages(url):
    """Visit multiple pages per site — like Hunter.io."""
    if not url or url == 'N/A':
        return ''
    clean = url if url.startswith('http') else f'https://{url}'
    clean = clean.split('?')[0].rstrip('/')
    subpages = ['','/contact','/contact-us','/contacts','/about','/about-us','/team','/our-team','/staff','/people','/leadership']
    all_html = ''
    for page in subpages:
        html = fetch_page(f'{clean}{page}')
        if html:
            all_html += html
        time.sleep(0.1)
    return all_html

# ── EMAIL ─────────────────────────────────────────────────────────────────────
def extract_email(html, domain):
    if not html:
        return 'N/A'
    emails = re.findall(r'[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}', html)
    blocked = ['example','domain','email','user','test','placeholder','png','jpg','gif','svg',
               'wix','sentry','wordpress','schema','w3.org','googleapis','cloudflare']
    domain_clean = re.sub(r'https?://(www\.)?','',str(domain or '')).split('/')[0].split('?')[0]
    priority = ['sales@','info@','contact@','hello@','support@','admin@','office@','team@']
    clean_emails = [e.lower() for e in emails if not any(b in e.lower() for b in blocked)]
    if not clean_emails:
        return 'N/A'
    for prefix in priority:
        for e in clean_emails:
            if e.startswith(prefix) and domain_clean in e:
                return e
    for e in clean_emails:
        if domain_clean in e:
            return e
    for prefix in priority:
        for e in clean_emails:
            if e.startswith(prefix):
                return e
    return clean_emails[0]

# ── SIGNALS ───────────────────────────────────────────────────────────────────
def extract_signals(html, url):
    if not html:
        return {
            'email':'N/A','services':'N/A','booking_online':'N/A',
            'insurance':'N/A','social_media':'N/A','tech_signals':'N/A',
        }

    soup = BeautifulSoup(html, 'lxml')
    text = soup.get_text(separator=' ', strip=True).lower()
    full_html = html.lower()

    # Email — from all pages
    all_html = fetch_all_pages(url)
    email = extract_email(all_html or html, url)

    # Services
    service_map = {
        'General Dentistry':['general dentistry','general dental','family dentistry','routine cleaning','checkup','examination'],
        'Cosmetic':         ['cosmetic dentistry','teeth whitening','veneers','smile makeover','aesthetic dentistry'],
        'Implants':         ['dental implants','implant dentistry','tooth implant','implant restoration'],
        'Orthodontics':     ['orthodontics','braces','invisalign','clear aligners','teeth straightening'],
        'Periodontics':     ['periodontics','gum disease','gum treatment','periodontist'],
        'Endodontics':      ['root canal','endodontics','endodontist'],
        'Oral Surgery':     ['oral surgery','tooth extraction','wisdom teeth','oral surgeon'],
        'Pediatric':        ['pediatric dentistry','children dentistry','kids dentist','child dental'],
        'Emergency':        ['emergency dental','emergency dentist','same day','urgent dental'],
        'Prosthodontics':   ['dentures','crowns','bridges','prosthodontics','full mouth'],
    }
    found_services = [s for s,kws in service_map.items() if any(kw in text for kw in kws)]
    services = ', '.join(found_services[:4]) if found_services else 'N/A'

    # Online booking
    booking_signals = ['book online','book an appointment','schedule online','request appointment',
                       'online booking','book now','schedule now','appointment request',
                       'zocdoc','open dental','dentrix','eaglesoft']
    booking_online = 'Yes' if any(s in text for s in booking_signals) else 'N/A'

    # Insurance
    insurance_signals = ['insurance','delta dental','cigna','metlife','aetna','guardian',
                         'bluecross','blue cross','united healthcare','humana','we accept',
                         'in-network','in network','insurance accepted']
    insurance = 'Accepted' if any(s in text for s in insurance_signals) else 'N/A'

    # Social media — expanded
    social = []
    if 'facebook.com' in full_html:                      social.append('Facebook')
    if 'instagram.com' in full_html:                     social.append('Instagram')
    if 'twitter.com' in full_html or 'x.com' in full_html: social.append('Twitter/X')
    if 'linkedin.com' in full_html:                      social.append('LinkedIn')
    if 'youtube.com' in full_html:                       social.append('YouTube')
    if 'tiktok.com' in full_html:                        social.append('TikTok')
    if 'pinterest.com' in full_html:                     social.append('Pinterest')
    if 'wa.me' in full_html or 'whatsapp.com' in full_html: social.append('WhatsApp')
    if 'yelp.com' in full_html:                          social.append('Yelp')
    if 'google.com/maps' in full_html or 'goo.gl/maps' in full_html: social.append('Google Business')
    social_media = ', '.join(social) if social else 'N/A'

    # Tech signals
    tech = []
    if any(s in full_html for s in ['zocdoc','zoc-doc']): tech.append('ZocDoc')
    if 'teledentistry' in text or 'virtual consult' in text: tech.append('Teledentistry')
    if 'patient portal' in text: tech.append('Patient Portal')
    if any(s in text for s in ['3d imaging','cone beam','cbct','digital xray','digital x-ray']): tech.append('Digital Imaging')
    if 'same day crown' in text or 'cerec' in text: tech.append('Same-Day Crowns')
    if 'laser dentistry' in text or 'laser treatment' in text: tech.append('Laser Dentistry')
    tech_signals = ', '.join(tech) if tech else 'N/A'

    return {
        'email':          email,
        'services':       services,
        'booking_online': booking_online,
        'insurance':      insurance,
        'social_media':   social_media,
        'tech_signals':   tech_signals,
    }

# ── SCORING ───────────────────────────────────────────────────────────────────
def enrichment_score_boost(signals):
    boost = 0
    if signals['email'] != 'N/A':           boost += 8
    if signals['booking_online'] == 'Yes':  boost += 8
    if signals['insurance'] == 'Accepted':  boost += 6
    if signals['tech_signals'] != 'N/A':    boost += 5
    if signals['social_media'] != 'N/A':    boost += 3
    services = signals['services']
    if services != 'N/A':
        count = len(services.split(','))
        if count >= 4:   boost += 7
        elif count >= 2: boost += 4
    if 'Implants' in services:     boost += 5
    if 'Cosmetic' in services:     boost += 5
    if 'Orthodontics' in services: boost += 3
    return boost

def get_tier(score):
    if score >= 70: return 'High'
    if score >= 40: return 'Medium'
    return 'Low'

# ── MAIN ──────────────────────────────────────────────────────────────────────
def enrich():
    print(f'🔬 Starting enrichment for {CITY}')

    wb_src = openpyxl.load_workbook(INPUT)
    ws_src = wb_src[CITY]

    rows = []
    for r in range(3, ws_src.max_row+1):
        row = [ws_src.cell(r, c).value for c in range(1, ws_src.max_column+1)]
        if any(row):
            rows.append(row)

    print(f'   Loaded {len(rows)} clinics')

    enriched = []
    for i, row in enumerate(rows):
        name    = row[1]
        website = row[5]
        print(f'  [{i+1}/{len(rows)}] {name}')

        html, _ = fetch_website(website), None
        if callable(html):
            html = html[0]

        # Clean fetch
        html = fetch_website(website)
        signals = extract_signals(html, website)
        boost   = enrichment_score_boost(signals)

        orig_score = int(row[11] or 0)
        new_score  = min(orig_score + boost, 100)
        new_tier   = get_tier(new_score)

        enriched.append({'row': row, 'signals': signals, 'score': new_score, 'tier': new_tier})
        time.sleep(0.3)

    # ── EXCEL ─────────────────────────────────────────────────────────────────
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    high = sum(1 for e in enriched if e['tier']=='High')
    med  = sum(1 for e in enriched if e['tier']=='Medium')
    low  = sum(1 for e in enriched if e['tier']=='Low')
    email_count = sum(1 for e in enriched if e['signals']['email'] != 'N/A')
    enrich_rate = f'{int(sum(1 for e in enriched if e["signals"]["services"]!="N/A")/len(enriched)*100)}%' if enriched else 'N/A'

    # Summary tab
    ws_sum = wb_out.create_sheet('Summary', 0)
    ws_sum.sheet_view.showGridLines = False

    ws_sum.merge_cells('A1:H1')
    c = ws_sum.cell(row=1, column=1, value=f'{CITY.upper()} DENTAL CLINIC INTELLIGENCE — LEADFLOW (ENRICHED)')
    c.font = Font(bold=True, name='Arial', size=15, color='FFFFFF')
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws_sum.row_dimensions[1].height = 34

    ws_sum.merge_cells('A2:H2')
    c = ws_sum.cell(row=2, column=1, value=f'⚠ Sample Database — {len(enriched)} Records  |  AI Enriched  |  Prepared by LeadFlow  |  June 2026  |  For Presentation Purposes Only')
    c.font = Font(bold=True, name='Arial', size=10, color='7a5c00')
    c.fill = GOLD_FILL
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws_sum.row_dimensions[2].height = 22
    ws_sum.row_dimensions[3].height = 8

    sum_hdrs   = ['City','Total','High','Medium','Low','Category','Enrichment Rate','Email Hit Rate']
    sum_widths = [16,10,10,12,10,16,18,16]
    for ci,(h,w) in enumerate(zip(sum_hdrs,sum_widths),1):
        c = ws_sum.cell(row=4, column=ci, value=h)
        c.font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
        c.fill = SUBHEAD_FILL
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = BORDER
        ws_sum.column_dimensions[get_column_letter(ci)].width = w
    ws_sum.row_dimensions[4].height = 20

    for ci,v in enumerate([CITY,len(enriched),high,med,low,'Dental Clinics',enrich_rate,f'{int(email_count/len(enriched)*100)}%'],1):
        c = ws_sum.cell(row=5, column=ci, value=v)
        c.font = Font(name='Arial', size=10)
        c.fill = ALT_FILL
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = BORDER
    ws_sum.row_dimensions[5].height = 20

    for ci,v in enumerate(['TOTAL',len(enriched),high,med,low,'','',''],1):
        c = ws_sum.cell(row=6, column=ci, value=v)
        c.font = Font(bold=True, color='FFFFFF', name='Arial', size=11)
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = BORDER
    ws_sum.row_dimensions[6].height = 24

    # City tab
    ws = wb_out.create_sheet(CITY)
    ws.sheet_view.showGridLines = False

    headers = [
        '#','Clinic Name','Phone','Address','Neighbourhood','Website',
        'Rating','Reviews','Opening Hours','GPS Latitude','GPS Longitude',
        'Email *','Services *','Online Booking *','Insurance *','Social Media *','Tech Signals *',
        'Score','Tier'
    ]
    widths = [
        5,38,16,46,18,28,
        8,10,22,14,14,
        28,48,16,14,40,40,
        8,10
    ]

    ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
    c = ws.cell(row=1, column=1, value=f'{CITY} — Dental Clinic Intelligence (AI Enriched)  |  {len(enriched)} Records  |  ⚠ Sample Database  |  LeadFlow  |  June 2026')
    c.font = Font(bold=True, name='Arial', size=10, color='FFFFFF')
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 24

    ws.merge_cells(f'A2:{get_column_letter(len(headers))}2')
    c = ws.cell(row=2, column=1, value='* Fields marked with asterisk are AI-enriched via website analysis. N/A shown where unavailable or not detected.')
    c.font = Font(italic=True, name='Arial', size=9, color='5B3FA8')
    c.fill = PatternFill('solid', start_color='F3EEFF')
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 16

    for ci,(h,w) in enumerate(zip(headers,widths),1):
        c = ws.cell(row=3, column=ci, value=h)
        if '*' in str(h):
            c.font = Font(bold=True, color='5B3FA8', name='Arial', size=10)
            c.fill = PatternFill('solid', start_color='DDD0FF')
        else:
            c.font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
            c.fill = SUBHEAD_FILL
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[3].height = 22

    for ri,e in enumerate(enriched, 4):
        row     = e['row']
        signals = e['signals']
        rfill   = ALT_FILL if ri%2==0 else PatternFill()
        tier_val = e['tier']

        vals = [
            ri-3,
            row[1],   # name
            row[2],   # phone
            row[3],   # address
            row[4],   # neighbourhood
            row[5],   # website
            row[6],   # rating
            row[7],   # reviews
            row[8],   # hours
            row[9],   # lat
            row[10],  # lng
            signals['email'],
            signals['services'],
            signals['booking_online'],
            signals['insurance'],
            signals['social_media'],
            signals['tech_signals'],
            e['score'],
            e['tier'],
        ]

        for ci,v in enumerate(vals,1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font = Font(name='Arial', size=10, color='000000')
            if ci >= 12 and ci <= 17:
                c.fill = ENR_FILL if ri%2==0 else ENR_FILL2
            else:
                c.fill = rfill
            c.alignment = Alignment(vertical='center', wrap_text=True)
            c.border = BORDER

        tier_cell = ws.cell(row=ri, column=19)
        tier_cell.fill = HIGH_FILL if tier_val=='High' else (MED_FILL if tier_val=='Medium' else LOW_FILL)
        tier_cell.font = Font(name='Arial', size=10, bold=True, color='000000')
        ws.row_dimensions[ri].height = 28

    ws.freeze_panes = 'A4'

    wb_out.save(OUTPUT)
    print(f'\n✅ Saved: {OUTPUT}')
    print(f'   High: {high} | Medium: {med} | Low: {low}')
    print(f'   Enrichment rate: {enrich_rate}')
    print(f'   Email hit rate: {int(email_count/len(enriched)*100)}%')

if __name__ == '__main__':
    enrich()
